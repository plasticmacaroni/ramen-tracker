#!/usr/bin/env python3
"""
Ramen Rater Data Pipeline

Downloads The Big List xlsx from theramenrater.com, converts it to JSON,
and fetches product images via Bing image search.
Each step skips work already done.

Usage:
    python fetch_ramen_data.py            # Process all ramen
    python fetch_ramen_data.py --limit 10 # Process up to 10 ramen that still need images
"""

import json
import os
import random
import re
import sys
import time
import argparse
from pathlib import Path
from io import BytesIO

# If running outside the venv, add its site-packages so imports work
_venv = Path(__file__).resolve().parent.parent / ".venv"
if _venv.is_dir() and not (hasattr(sys, 'real_prefix') or sys.base_prefix != sys.prefix):
    _sp = _venv / "Lib" / "site-packages"
    if not _sp.is_dir():
        _py = f"python{sys.version_info.major}.{sys.version_info.minor}"
        _sp = _venv / "lib" / _py / "site-packages"
    if _sp.is_dir() and str(_sp) not in sys.path:
        sys.path.insert(0, str(_sp))

import requests
from openpyxl import load_workbook

import threading
import tkinter as tk

XLSX_URL = "https://www.theramenrater.com/wp-content/uploads/2025/11/11212025The-Ramen-Rater.xlsx"
ROOT_DIR = Path(__file__).resolve().parent.parent
DATA_DIR = ROOT_DIR / "data"
IMAGES_DIR = ROOT_DIR / "images" / "ramen"
BRAND_DIR = ROOT_DIR / "images" / "brand"
CACHE_DIR = ROOT_DIR / "tools" / ".cache"
XLSX_PATH = CACHE_DIR / "big-list.xlsx"
XLSX_ETAG_PATH = CACHE_DIR / ".big-list-etag"

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                  "(KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36",
}

def _load_typos():
    typo_path = Path(__file__).resolve().parent / "typos.json"
    if typo_path.exists():
        with open(typo_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {}

_TYPOS = _load_typos()
COUNTRY_TYPOS = {k.lower(): v for k, v in _TYPOS.get('country', {}).items()}
STYLE_TYPOS = {k.lower(): v for k, v in _TYPOS.get('style', {}).items()}
BRAND_TYPOS = {k.lower(): v for k, v in _TYPOS.get('brand', {}).items()}
TEXT_TYPOS = {k.lower(): v for k, v in _TYPOS.get('text', {}).items()}
RENAMES = {r['id']: r for r in _TYPOS.get('rename', [])}
EXCLUDES = set(_TYPOS.get('exclude', []))


def download_xlsx():
    """Download the xlsx, using ETag to avoid redundant downloads."""
    CACHE_DIR.mkdir(parents=True, exist_ok=True)
    headers = dict(HEADERS)

    if XLSX_PATH.exists() and XLSX_ETAG_PATH.exists():
        saved_etag = XLSX_ETAG_PATH.read_text().strip()
        headers['If-None-Match'] = saved_etag

    print(f"Checking xlsx at {XLSX_URL}...")
    resp = requests.get(XLSX_URL, headers=headers, timeout=60)

    if resp.status_code == 304:
        print("  xlsx unchanged, skipping download.")
        return

    resp.raise_for_status()
    XLSX_PATH.write_bytes(resp.content)
    print(f"  Downloaded ({len(resp.content) / 1024:.0f} KB)")

    etag = resp.headers.get('ETag')
    if etag:
        XLSX_ETAG_PATH.write_text(etag)


def parse_xlsx():
    print(f"Parsing {XLSX_PATH}...")
    wb = load_workbook(XLSX_PATH, read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        print("  ERROR: Spreadsheet is empty")
        return []

    header_idx = 0
    for i, row in enumerate(rows[:5]):
        cells = [str(c).strip().lower() if c else '' for c in row]
        if any('review' in c for c in cells):
            header_idx = i
            break

    headers_raw = [str(c).strip() if c else '' for c in rows[header_idx]]
    print(f"  Headers (row {header_idx}): {headers_raw}")

    col_map = {}
    for i, h in enumerate(headers_raw):
        hl = h.lower()
        if 'review' in hl and '#' in hl:
            col_map['id'] = i
        elif hl == 'brand':
            col_map['brand'] = i
        elif hl == 'variety':
            col_map['variety'] = i
        elif hl == 'style':
            col_map['style'] = i
        elif hl == 'country':
            col_map['country'] = i
        elif hl == 'stars':
            col_map['stars'] = i

    required = ['id', 'brand', 'variety']
    missing = [k for k in required if k not in col_map]
    if missing:
        print(f"  ERROR: Missing columns: {missing}")
        return []

    ramen_list = []
    for row in rows[header_idx + 1:]:
        try:
            raw_id = row[col_map['id']]
            if raw_id is None:
                continue
            review_id = int(float(str(raw_id)))
        except (ValueError, TypeError):
            continue

        if review_id in EXCLUDES:
            continue

        brand = str(row[col_map.get('brand', 0)] or '').strip()
        brand = BRAND_TYPOS.get(brand.lower(), brand)
        variety = str(row[col_map.get('variety', 0)] or '').strip()
        for typo, fix in TEXT_TYPOS.items():
            escaped = re.escape(typo)
            if re.match(r'\w', typo):
                escaped = r'\b' + escaped
            if re.search(r'\w$', typo):
                escaped = escaped + r'\b'
            variety = re.sub(escaped, fix, variety, flags=re.IGNORECASE)
        if not variety:
            continue

        style = str(row[col_map.get('style', 0)] or '').strip() if 'style' in col_map else ''
        style = STYLE_TYPOS.get(style.lower(), style)
        country = str(row[col_map.get('country', 0)] or '').strip() if 'country' in col_map else ''
        country = COUNTRY_TYPOS.get(country.lower(), country)

        stars = None
        if 'stars' in col_map:
            raw_stars = row[col_map['stars']]
            if raw_stars is not None:
                try:
                    stars = float(raw_stars)
                except (ValueError, TypeError):
                    pass

        rename = RENAMES.get(review_id)
        if rename:
            if 'replace_variety' in rename:
                variety = rename['replace_variety']
            if 'replace_brand' in rename:
                brand = rename['replace_brand']
            if 'replace_style' in rename:
                style = rename['replace_style']
            if 'replace_country' in rename:
                country = rename['replace_country']
            if 'replace_stars' in rename:
                stars = float(rename['replace_stars'])

        ramen_list.append({
            'id': review_id,
            'brand': brand,
            'variety': variety,
            'style': style,
            'country': country,
            'stars': stars,
        })

    wb.close()
    print(f"  Parsed {len(ramen_list)} ramen entries")
    return ramen_list


def save_json(ramen_list):
    DATA_DIR.mkdir(exist_ok=True)
    ramen_list.sort(key=lambda r: r.get("id", 0))
    out_path = DATA_DIR / "ramen.json"
    with open(out_path, 'w', encoding='utf-8') as f:
        json.dump(ramen_list, f, ensure_ascii=False)
    print(f"  Saved {out_path} ({out_path.stat().st_size / 1024:.0f} KB)")


def cleanup_excluded():
    """Remove images and urls for excluded IDs."""
    ramen_ids = set()
    ramen_path = DATA_DIR / "ramen.json"
    if ramen_path.exists():
        with open(ramen_path, 'r', encoding='utf-8') as f:
            ramen_ids = {str(r['id']) for r in json.load(f)}

    removed_images = 0
    if IMAGES_DIR.exists():
        for f in IMAGES_DIR.glob("*.webp"):
            if f.stem.isdigit() and f.stem not in ramen_ids:
                f.unlink()
                removed_images += 1

    removed_urls = 0
    urls_path = DATA_DIR / "urls.json"
    if urls_path.exists():
        with open(urls_path, 'r', encoding='utf-8') as f:
            urls = json.load(f)
        before = len(urls)
        urls = {k: v for k, v in urls.items() if k in ramen_ids}
        removed_urls = before - len(urls)
        if removed_urls:
            with open(urls_path, 'w', encoding='utf-8') as f:
                json.dump(urls, f, indent=2, ensure_ascii=False)
                f.write('\n')

    if removed_images or removed_urls:
        print(f"  Cleaned up excluded entries: {removed_images} images, {removed_urls} URLs removed")


POPULARITY_PATH = DATA_DIR / "popularity.json"
REFRESHED_PATH = DATA_DIR / "popularity_refreshed.json"
def load_popularity():
    """Load {id: count} from data/popularity.json. Returns dict with int keys."""
    if POPULARITY_PATH.exists():
        with open(POPULARITY_PATH, 'r', encoding='utf-8') as f:
            raw = json.load(f)
        return {int(k): v for k, v in raw.items()}
    return {}


def save_popularity(pop_map):
    """Write {id: count} to data/popularity.json (string keys for JSON compat)."""
    DATA_DIR.mkdir(exist_ok=True)
    with open(POPULARITY_PATH, 'w', encoding='utf-8') as f:
        json.dump({str(k): v for k, v in sorted(pop_map.items())}, f, ensure_ascii=False)
    print(f"  Saved {POPULARITY_PATH} ({len(pop_map)} entries)")


def _load_refreshed():
    if REFRESHED_PATH.exists():
        with open(REFRESHED_PATH, 'r', encoding='utf-8') as f:
            return set(json.load(f))
    return set()


def _save_refreshed(ids):
    with open(REFRESHED_PATH, 'w', encoding='utf-8') as f:
        json.dump(sorted(ids), f)


def next_refresh_candidate(pop_map):
    """Return the ramen ID with the highest popularity that hasn't been refreshed yet, or None."""
    refreshed = _load_refreshed()
    best_id, best_score = None, -1
    for rid, count in pop_map.items():
        if rid not in refreshed and count > best_score:
            best_id, best_score = rid, count
    return best_id


def mark_popularity_refreshed(rid):
    """Mark an ID as refreshed after its new score has been saved."""
    refreshed = _load_refreshed()
    refreshed.add(rid)
    _save_refreshed(refreshed)


def count_unrefreshed(pop_map):
    """Count how many popularity entries haven't been refreshed yet."""
    refreshed = _load_refreshed()
    return sum(1 for rid in pop_map if rid not in refreshed)


BROWSER_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                  "(KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
}

def _quote_brand(brand):
    """Quote each slash-separated part of a brand for search queries."""
    if '/' not in brand:
        return f'"{brand}"'
    parts = [p.strip() for p in brand.split('/') if p.strip()]
    return ' '.join(f'"{p}"' for p in parts)


def _search_bing(query):
    """Scrape Bing image search results. Returns list of image URLs.
    Retries with back-off on 403/429 rate-limiting responses."""
    from urllib.parse import quote_plus
    url = f"https://www.bing.com/images/search?q={quote_plus(query)}&form=HDRSC2&first=1"
    for attempt in range(3):
        try:
            resp = requests.get(url, headers=BROWSER_HEADERS, timeout=15)
            if resp.status_code in (403, 429):
                wait = 2 ** (attempt + 1)
                print(f"      Bing: {resp.status_code} — retrying in {wait}s ({attempt+1}/3)")
                time.sleep(wait)
                continue
            resp.raise_for_status()
            urls = re.findall(r'murl&quot;:&quot;(https?://[^&]+?)&quot;', resp.text)
            urls = [u for u in urls if not u.endswith('.svg')]
            return urls[:5]
        except requests.RequestException as e:
            print(f"      Bing: {type(e).__name__}: {e}")
            return []
    print(f"      Bing: still blocked after 3 retries, skipping")
    return []


UBLOCK_ID = "ddkjiahejlhfcafbddmgiahcphecmpfh"
UBLOCK_DIR = CACHE_DIR / "ublock"
UBLOCK_VERSION_FILE = CACHE_DIR / "ublock-version"
PW_VERSION_FILE = CACHE_DIR / "playwright-version"


def _ensure_ublock():
    """Download/update uBlock Origin Lite if needed. Returns path to unpacked extension."""
    import zipfile

    crx_url = (
        f"https://clients2.google.com/service/update2/crx"
        f"?response=redirect&prodversion=130.0&acceptformat=crx2,crx3"
        f"&x=id%3D{UBLOCK_ID}%26uc"
    )

    manifest_path = UBLOCK_DIR / "manifest.json"
    local_ver = ""
    if UBLOCK_VERSION_FILE.exists():
        local_ver = UBLOCK_VERSION_FILE.read_text().strip()

    # Check for update via HEAD request ETag
    try:
        head = requests.head(crx_url, headers=BROWSER_HEADERS, timeout=10, allow_redirects=True)
        remote_tag = head.headers.get("ETag", "") or head.headers.get("Last-Modified", "")
    except Exception:
        remote_tag = ""

    if manifest_path.exists() and local_ver and local_ver == remote_tag and remote_tag:
        print(f"  uBlock Origin Lite: up to date")
        return str(UBLOCK_DIR)

    print(f"  Downloading uBlock Origin Lite...")
    resp = requests.get(crx_url, headers=BROWSER_HEADERS, timeout=30, allow_redirects=True)
    resp.raise_for_status()
    crx_data = resp.content

    # CRX3 format: magic(4) + version(4) + header_len(4) + header + ZIP
    # Find the ZIP start (PK signature)
    zip_start = crx_data.find(b'PK\x03\x04')
    if zip_start < 0:
        raise RuntimeError("Could not find ZIP data in CRX file")

    if UBLOCK_DIR.exists():
        import shutil
        shutil.rmtree(UBLOCK_DIR)
    UBLOCK_DIR.mkdir(parents=True)

    with zipfile.ZipFile(BytesIO(crx_data[zip_start:])) as zf:
        zf.extractall(UBLOCK_DIR)

    version_tag = remote_tag or "downloaded"
    UBLOCK_VERSION_FILE.write_text(version_tag)

    # Read actual extension version from manifest
    if manifest_path.exists():
        import json as _json
        mf = _json.loads(manifest_path.read_text(encoding="utf-8"))
        print(f"  uBlock Origin Lite v{mf.get('version', '?')} installed")
    else:
        print(f"  uBlock Origin Lite installed")

    return str(UBLOCK_DIR)


def _ensure_playwright_browser():
    """Run 'playwright install chromium' if the Playwright version has changed."""
    import playwright
    current_ver = getattr(playwright, '__version__', '')

    saved_ver = ""
    if PW_VERSION_FILE.exists():
        saved_ver = PW_VERSION_FILE.read_text().strip()

    if current_ver and current_ver == saved_ver:
        return

    import subprocess
    print(f"  Installing Playwright Chromium (v{current_ver})...")
    subprocess.run([sys.executable, "-m", "playwright", "install", "chromium"],
                   check=True, capture_output=True)
    PW_VERSION_FILE.write_text(current_ver)
    print(f"  Playwright Chromium ready")


class ScraperControlPanel:
    """Floating tkinter window for live control during scraping.
    
    Tkinter must own the main thread, so the scraping work runs in a
    background thread while this panel runs mainloop() on main.
    """

    def __init__(self, ramen_list):
        self._ramen_list = ramen_list
        self.engine = "google"
        self._lock = threading.Lock()
        self._single_queue = []
        self._bulk_paused = False
        self._wake = threading.Event()
        self._match_ids = []
        self.shutting_down = False

        self._root = tk.Tk()
        root = self._root
        root.title("Ramen Scraper")
        root.attributes("-topmost", True)
        root.resizable(True, True)
        root.geometry("620x750")
        root.minsize(460, 600)
        root.configure(bg="#1a1a2e")

        tk.Label(root, text="Search Engine", font=("Segoe UI", 10, "bold"),
                 fg="#e0e0e0", bg="#1a1a2e").pack(pady=(10, 2))

        self._engine_var = tk.StringVar(value="google")
        frame = tk.Frame(root, bg="#1a1a2e")
        frame.pack()
        for val, label in [("google", "Google"), ("bing", "Bing")]:
            tk.Radiobutton(frame, text=label, variable=self._engine_var, value=val,
                           command=self._on_engine_change,
                           fg="#e0e0e0", bg="#1a1a2e", selectcolor="#16213e",
                           activebackground="#1a1a2e", activeforeground="#f7d354",
                           font=("Segoe UI", 10)).pack(side=tk.LEFT, padx=10)

        tk.Label(root, text="Scrape one ramen (pauses bulk queue)", font=("Segoe UI", 9, "bold"),
                 fg="#e0e0e0", bg="#1a1a2e").pack(pady=(12, 2))
        tk.Label(root, text="Fuzzy search — pick a row, then queue scrape (or double-click row)",
                 font=("Segoe UI", 8), fg="#888", bg="#1a1a2e").pack(pady=(0, 2))

        self._single_var = tk.StringVar()
        ent = tk.Entry(root, textvariable=self._single_var, font=("Segoe UI", 10),
                       width=52, bg="#16213e", fg="#e0e0e0", insertbackground="#e0e0e0")
        ent.pack(pady=(0, 4), padx=10)
        ent.bind("<Return>", lambda e: self._on_find_matches())

        find_row = tk.Frame(root, bg="#1a1a2e")
        find_row.pack(pady=(0, 6))
        tk.Button(find_row, text="Find matches", command=self._on_find_matches,
                  font=("Segoe UI", 9), bg="#16213e", fg="#e0e0e0",
                  activebackground="#0f3460", activeforeground="#f7d354").pack(side=tk.LEFT, padx=4)

        list_frame = tk.Frame(root, bg="#1a1a2e")
        list_frame.pack(pady=(0, 6), padx=10, fill=tk.BOTH, expand=True)
        yscroll = tk.Scrollbar(list_frame, orient=tk.VERTICAL)
        yscroll.pack(side=tk.RIGHT, fill=tk.Y)
        xscroll = tk.Scrollbar(list_frame, orient=tk.HORIZONTAL)
        xscroll.pack(side=tk.BOTTOM, fill=tk.X)
        self._results_list = tk.Listbox(
            list_frame, height=12, font=("TkFixedFont", 9),
            bg="#16213e", fg="#e0e0e0", selectbackground="#0f3460",
            selectforeground="#f7d354",
            yscrollcommand=yscroll.set, xscrollcommand=xscroll.set,
            exportselection=False, activestyle="dotbox",
        )
        self._results_list.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        yscroll.config(command=self._results_list.yview)
        xscroll.config(command=self._results_list.xview)
        self._results_list.bind("<Double-Button-1>", lambda e: self._on_queue_scrape_selection())

        btn_row = tk.Frame(root, bg="#1a1a2e")
        btn_row.pack(pady=(0, 6))
        tk.Button(btn_row, text="Queue scrape (selected)", command=self._on_queue_scrape_selection,
                  font=("Segoe UI", 9), bg="#16213e", fg="#e0e0e0",
                  activebackground="#0f3460", activeforeground="#f7d354").pack(side=tk.LEFT, padx=4)
        tk.Button(btn_row, text="Resume all scraping", command=self._on_resume_all,
                  font=("Segoe UI", 9), bg="#16213e", fg="#e0e0e0",
                  activebackground="#0f3460", activeforeground="#f7d354").pack(side=tk.LEFT, padx=4)

        tk.Frame(root, bg="#333", height=1).pack(fill=tk.X, padx=10, pady=(6, 4))

        stats_frame = tk.Frame(root, bg="#1a1a2e")
        stats_frame.pack(fill=tk.X, padx=10, pady=(0, 2))

        left_stats = tk.Frame(stats_frame, bg="#1a1a2e")
        left_stats.pack(side=tk.LEFT, fill=tk.X, expand=True)

        self._captcha_count = 0
        self._since_captcha_var = tk.StringVar(value="Since CAPTCHA: 0")
        tk.Label(left_stats, textvariable=self._since_captcha_var,
                 font=("Segoe UI", 10, "bold"), fg="#2ecc71", bg="#1a1a2e",
                 anchor="w").pack(fill=tk.X)

        self._total_scored_var = tk.StringVar(value="Scored: 0")
        tk.Label(left_stats, textvariable=self._total_scored_var,
                 font=("Segoe UI", 9), fg="#a0a0a0", bg="#1a1a2e",
                 anchor="w").pack(fill=tk.X)

        tk.Frame(root, bg="#333", height=1).pack(fill=tk.X, padx=10, pady=(4, 2))

        tk.Label(root, text="Recent Results", font=("Segoe UI", 9, "bold"),
                 fg="#e0e0e0", bg="#1a1a2e", anchor="w").pack(fill=tk.X, padx=10, pady=(2, 0))

        recent_frame = tk.Frame(root, bg="#1a1a2e")
        recent_frame.pack(fill=tk.X, padx=10, pady=(0, 4))
        self._recent_list = tk.Listbox(
            recent_frame, height=6, font=("TkFixedFont", 8),
            bg="#16213e", fg="#e0e0e0", selectbackground="#0f3460",
            selectforeground="#f7d354", exportselection=False, activestyle="none",
        )
        self._recent_list.pack(fill=tk.X, expand=True)

        tk.Frame(root, bg="#333", height=1).pack(fill=tk.X, padx=10, pady=(2, 4))

        self._progress_var = tk.StringVar(value="Starting...")
        tk.Label(root, textvariable=self._progress_var, font=("Segoe UI", 9),
                 fg="#a0a0a0", bg="#1a1a2e").pack(pady=(2, 2))

        self._status_var = tk.StringVar(value="")
        tk.Label(root, textvariable=self._status_var, font=("Segoe UI", 8),
                 fg="#666", bg="#1a1a2e", wraplength=430).pack(pady=(0, 6))

        root.protocol("WM_DELETE_WINDOW", self._on_close)

    def _on_close(self):
        self.shutting_down = True
        self._wake.set()
        try:
            self._root.destroy()
        except Exception:
            pass

    def _on_engine_change(self):
        self.engine = self._engine_var.get()

    def _on_find_matches(self):
        text = (self._single_var.get() or "").strip()
        if not text:
            self.set_status("Type a few letters, a review #, or brand words")
            return
        matches = _fuzzy_rank_ramen(text, self._ramen_list, limit=50)
        self._results_list.delete(0, tk.END)
        self._match_ids.clear()
        if not matches:
            self.set_status("No matches — try different words or review #")
            return
        for r in matches:
            brand = r.get("brand") or ""
            variety = r.get("variety") or ""
            line = f"#{r['id']:5d}  {brand} — {variety}"
            self._results_list.insert(tk.END, line)
            self._match_ids.append(r["id"])
        self.set_status(f"{len(matches)} matches — select one, Queue scrape (or double-click)")

    def _on_queue_scrape_selection(self):
        sel = self._results_list.curselection()
        if not sel:
            self.set_status("Select a row in the list first (Find matches, then click a line)")
            return
        rid = self._match_ids[sel[0]]
        with self._lock:
            self._single_queue.append(rid)
            self._bulk_paused = True
        self._wake.set()
        self.set_status(f"Queued #{rid} — runs after current step, then bulk stays paused")

    def _on_resume_all(self):
        with self._lock:
            self._bulk_paused = False
        self._wake.set()
        self.set_status("Resuming bulk queue…")

    def set_progress(self, text):
        if self._root:
            self._root.after(0, lambda: self._progress_var.set(text))

    def set_status(self, text):
        if self._root:
            self._root.after(0, lambda: self._status_var.set(text))

    def record_captcha(self):
        self._captcha_count += 1
        def _update():
            self._since_captcha_var.set(f"Since CAPTCHA: 0  (total: {self._captcha_count})")
        if self._root:
            self._root.after(0, _update)

    def record_success(self, rid, brand, variety, count):
        if self._root:
            self._root.after(0, lambda: self._add_recent(rid, brand, variety, count))

    def record_error(self, rid, brand, variety, reason="NO RESULTS"):
        if self._root:
            self._root.after(0, lambda: self._add_recent_error(rid, brand, variety, reason))

    def _add_recent_error(self, rid, brand, variety, reason):
        line = f"#{rid:<5d} *** {reason} ***  {brand} — {variety}"
        self._recent_list.insert(0, line)
        self._recent_list.itemconfig(0, fg="#ff4444", selectforeground="#ff6666")
        if self._recent_list.size() > 50:
            self._recent_list.delete(50, tk.END)

    def _add_recent(self, rid, brand, variety, count):
        line = f"#{rid:<5d} {count:>12,}  {brand} — {variety}"
        self._recent_list.insert(0, line)
        if self._recent_list.size() > 50:
            self._recent_list.delete(50, tk.END)
        cur = self._since_captcha_var.get()
        try:
            parts = cur.split(":")
            rest = parts[1].strip()
            if "(" in rest:
                num = int(rest.split("(")[0].strip())
                total_part = rest.split("(")[1].rstrip(")")
                self._since_captcha_var.set(f"Since CAPTCHA: {num + 1}  ({total_part})")
            else:
                num = int(rest)
                self._since_captcha_var.set(f"Since CAPTCHA: {num + 1}")
        except Exception:
            pass

    def update_scored_total(self, scored, total):
        if self._root:
            self._root.after(0, lambda: self._total_scored_var.set(
                f"Scored: {scored} / {total} remaining"))

    def run_with_panel(self, fn, *args, **kwargs):
        """Run fn in a background thread while tkinter mainloop owns the main thread."""
        result = [None]
        error = [None]

        def _worker():
            try:
                result[0] = fn(*args, **kwargs)
            except KeyboardInterrupt:
                self.shutting_down = True
            except Exception as e:
                error[0] = e
            finally:
                try:
                    self._root.after(0, self._root.destroy)
                except Exception:
                    pass

        def _signal_check():
            """Periodic no-op that lets Python's signal handler fire inside mainloop."""
            if self.shutting_down:
                self._root.destroy()
                return
            try:
                self._root.after(200, _signal_check)
            except Exception:
                pass

        import signal
        prev_handler = signal.getsignal(signal.SIGINT)

        def _ctrl_c(sig, frame):
            print("\n\nCtrl+C — shutting down…")
            self.shutting_down = True
            self._wake.set()
            try:
                self._root.after(0, self._root.destroy)
            except Exception:
                pass

        signal.signal(signal.SIGINT, _ctrl_c)

        t = threading.Thread(target=_worker, daemon=True)
        t.start()
        self._root.after(200, _signal_check)
        try:
            self._root.mainloop()
        except Exception:
            pass
        finally:
            signal.signal(signal.SIGINT, prev_handler)
        t.join(timeout=3)
        if self.shutting_down:
            raise KeyboardInterrupt
        if error[0]:
            raise error[0]
        return result[0]

    def destroy(self):
        if self._root:
            try:
                self._root.destroy()
            except Exception:
                pass


def _create_popularity_browser():
    """Launch a Playwright browser with uBlock Origin Lite. Returns (playwright, context, page)."""
    from playwright.sync_api import sync_playwright

    _ensure_playwright_browser()
    ext_path = _ensure_ublock()

    pw = sync_playwright().start()
    user_data = str(CACHE_DIR / "pw-profile")
    context = pw.chromium.launch_persistent_context(
        user_data,
        headless=False,
        args=[
            f"--disable-extensions-except={ext_path}",
            f"--load-extension={ext_path}",
            "--disable-blink-features=AutomationControlled",
        ],
        user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                   "(KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36",
    )
    page = context.new_page()
    return pw, context, page


DEBUG_LOG = CACHE_DIR / "popularity-debug.log"


def _debug(msg):
    CACHE_DIR.mkdir(parents=True, exist_ok=True)
    with open(DEBUG_LOG, 'a', encoding='utf-8') as f:
        f.write(f"[{time.strftime('%H:%M:%S')}] {msg}\n")


_on_captcha = None

def _bing_web_result_count(page, query, brand):
    """Get estimated result count from Bing web search via Playwright. Returns int."""
    from urllib.parse import quote_plus
    url = f"https://www.bing.com/search?q={quote_plus(query)}"
    _debug(f"QUERY: {query}")
    _debug(f"  URL: {url}")
    _debug(f"  BRAND: {brand}")
    try:
        page.goto(url, wait_until="load", timeout=20000)

        el = page.wait_for_selector(".sb_count", timeout=5000)
        if not el:
            raise TimeoutError()
    except Exception:
        print(f"      Waiting for CAPTCHA — solve it in the browser, will resume automatically...")
        if _on_captcha:
            _on_captcha()
        _debug(f"  .sb_count not found in 5s, waiting indefinitely (CAPTCHA?)")
        try:
            page.wait_for_selector(".sb_count", timeout=0)
        except Exception as e:
            _debug(f"  FAILED: {type(e).__name__}: {e}")
            _debug(f"  PAGE TITLE: {page.title()}")
            _debug(f"  PAGE URL: {page.url}")
            body_text = page.inner_text("body")[:500]
            _debug(f"  BODY START: {body_text}")
            return 0

    try:
        text = page.inner_text(".sb_count")
        _debug(f"  .sb_count raw text: {repr(text)}")
        match = re.search(r'([\d,\.]+)', text)
        if match:
            count = int(match.group(1).replace(',', '').replace('.', ''))
            _debug(f"  PARSED: {count}")
            _maybe_click_result(page, "bing")
            return count
        else:
            _debug(f"  NO MATCH in: {repr(text)}")
    except Exception as e:
        _debug(f"  ERROR reading .sb_count: {type(e).__name__}: {e}")
    return 0


def _maybe_click_result(page, engine):
    """40% chance to click a random organic result, dwell 3-12s, then go back.
    Prefers The Ramen Rater links if present."""
    if random.random() > 0.05:
        return
    try:
        if engine == "google":
            links = page.query_selector_all("#search a h3")
        else:
            links = page.query_selector_all("#b_results h2 a")
        if not links:
            return
        # Prefer The Ramen Rater links
        rr_links = []
        for link in links:
            href = link.evaluate('el => el.closest("a")?.href || ""')
            if 'theramenrater' in href.lower():
                rr_links.append(link)
        pick = random.choice(rr_links) if rr_links else random.choice(links)
        pick.scroll_into_view_if_needed()
        time.sleep(random.uniform(0.3, 1.0))
        pick.click()
        dwell = random.uniform(3, 12)
        _debug(f"  DWELL: clicked random result, staying {dwell:.1f}s")
        time.sleep(dwell)
        page.go_back(wait_until="load", timeout=10000)
    except Exception as e:
        _debug(f"  DWELL error: {type(e).__name__}: {e}")


def _google_web_result_count(page, query, brand):
    """Get estimated result count from Google web search via Playwright. Returns int."""
    from urllib.parse import quote_plus
    url = f"https://www.google.com/search?q={quote_plus(query)}"
    _debug(f"GOOGLE QUERY: {query}")
    _debug(f"  URL: {url}")
    _debug(f"  BRAND: {brand}")
    try:
        page.goto(url, wait_until="domcontentloaded", timeout=20000)
        page.wait_for_selector("#search", timeout=5000)
    except Exception:
        print(f"      Waiting for CAPTCHA — solve it in the browser, will resume automatically...")
        if _on_captcha:
            _on_captcha()
        _debug(f"  #search not found in 5s, waiting indefinitely (CAPTCHA?)")
        try:
            page.wait_for_selector("#search", timeout=0)
        except Exception as e:
            _debug(f"  FAILED: {type(e).__name__}: {e}")
            _debug(f"  PAGE TITLE: {page.title()}")
            _debug(f"  PAGE URL: {page.url}")
            body_text = page.inner_text("body")[:500]
            _debug(f"  BODY START: {body_text}")
            return 0

    # Wait for actual search results to render inside #search (not just the container)
    try:
        page.wait_for_load_state("networkidle", timeout=10000)
    except Exception:
        _debug(f"  networkidle timed out, proceeding anyway")

    # Wait for at least one result heading or the "About" result-stats to appear
    try:
        page.wait_for_selector("#result-stats, #search h3, #rso", timeout=5000)
        _debug(f"  Search results content detected")
    except Exception:
        _debug(f"  No result headings found after waiting")

    time.sleep(random.uniform(1, 3))

    # Click "Tools" to open the About section with result count
    try:
        tools_btn = page.query_selector("#hdtb-tls") or page.query_selector("div.hdtb-mitem >> text=Tools")
        if tools_btn:
            tools_btn.click()
            _debug(f"  Clicked Tools button")
        else:
            _debug(f"  Tools button not found, trying #result-stats directly")
    except Exception as e:
        _debug(f"  Tools click issue: {type(e).__name__}: {e}")

    # Wait for the "About X results" section to actually populate
    try:
        page.wait_for_function(
            '() => { const el = document.getElementById("result-stats"); return el && el.innerText.length > 0; }',
            timeout=5000
        )
        _debug(f"  #result-stats populated")
    except Exception as e:
        _debug(f"  #result-stats did not populate: {type(e).__name__}: {e}")

    try:
        text = page.inner_text("#result-stats")
        _debug(f"  #result-stats raw text: {repr(text)}")
        match = re.search(r'([\d,\.]+)\s+result', text)
        if match:
            count = int(match.group(1).replace(',', '').replace('.', ''))
            _debug(f"  PARSED: {count}")
            _maybe_click_result(page, "google")
            return count
        else:
            _debug(f"  NO MATCH in: {repr(text)}")
    except Exception as e:
        _debug(f"  ERROR reading #result-stats: {type(e).__name__}: {e}")
    return 0


def _download_image(img_url, out_path, has_pillow, Image):
    """Download a single image URL and save as compressed WebP. Returns True on success."""
    try:
        resp = requests.get(img_url, headers={**BROWSER_HEADERS, "Accept": "image/webp,image/*,*/*;q=0.8"}, timeout=15)
        if resp.status_code in (403, 401):
            print(f"      Skipped ({resp.status_code} Forbidden)")
            return False
        resp.raise_for_status()

        content_type = resp.headers.get('Content-Type', '')
        if 'html' in content_type:
            print(f"      Skipped (got HTML instead of image)")
            return False

        if len(resp.content) < 1000:
            print(f"      Skipped (only {len(resp.content)} bytes)")
            return False

        if has_pillow:
            from PIL import ImageOps
            img = Image.open(BytesIO(resp.content))
            img = ImageOps.exif_transpose(img)
            img = img.convert('RGB')
            if img.width > MAX_WIDTH:
                ratio = MAX_WIDTH / img.width
                img = img.resize((MAX_WIDTH, int(img.height * ratio)), Image.LANCZOS)
            img.info.pop('exif', None)
            img.save(out_path, 'WEBP', quality=WEBP_QUALITY)
        else:
            out_path.write_bytes(resp.content)

        return True

    except requests.RequestException as e:
        print(f"      Download failed: {type(e).__name__}: {e}")
    except Exception as e:
        print(f"      Image processing failed: {type(e).__name__}: {e}")
    return False


MAX_WIDTH = 400
WEBP_QUALITY = 75
MAX_FILE_SIZE = 80_000  # 80KB threshold for recompression


def _recompress_single(path, Image):
    """Recompress a single image if it exceeds size/width limits."""
    if not path.exists() or Image is None:
        return
    try:
        from PIL import ImageOps
        file_size = path.stat().st_size
        img = Image.open(path)
        reasons = []

        exif_orientation = None
        try:
            exif_orientation = img.getexif().get(0x0112, 1)
        except Exception:
            pass
        if exif_orientation and exif_orientation != 1:
            img = ImageOps.exif_transpose(img)
            reasons.append(f"EXIF rotated")

        if file_size > MAX_FILE_SIZE:
            reasons.append(f"too large ({file_size//1024}KB)")
        if img.width > MAX_WIDTH:
            reasons.append(f"too wide ({img.width}px)")

        if not reasons:
            return

        img = img.convert('RGB')
        if img.width > MAX_WIDTH:
            ratio = MAX_WIDTH / img.width
            img = img.resize((MAX_WIDTH, int(img.height * ratio)), Image.LANCZOS)

        fmt = path.suffix.lstrip('.').upper()
        quality_kw = {'quality': WEBP_QUALITY} if fmt == 'WEBP' else {'optimize': True}
        img.save(path, fmt, **quality_kw)
        new_size = path.stat().st_size
        print(f"      Recompressed: {', '.join(reasons)} -> {new_size//1024}KB ({img.width}x{img.height})")
    except Exception as e:
        print(f"      Recompress error: {e}")


def _recompress_scan(Image):
    """Scan all ramen and brand images, recompress any that need it."""
    for directory, pattern in [(IMAGES_DIR, "*.webp"), (BRAND_DIR, "*.png")]:
        if not directory.exists():
            continue
        for img_file in directory.glob(pattern):
            try:
                _recompress_single(img_file, Image)
            except Exception:
                pass


def _recompress_dir(directory, patterns, fmt, quality, Image):
    """Recompress images in a directory that exceed size/width limits. Returns count."""
    if not directory.exists():
        return 0

    files = []
    for pat in patterns:
        files.extend(directory.glob(pat))
    if not files:
        return 0

    from PIL import ImageOps

    recompressed = 0
    for f in files:
        try:
            file_size = f.stat().st_size
            img = Image.open(f)
            reasons = []

            # Check EXIF orientation (tag 0x0112; value 1 = normal)
            exif_orientation = None
            try:
                exif = img.getexif()
                exif_orientation = exif.get(0x0112, 1)
            except Exception:
                pass
            if exif_orientation and exif_orientation != 1:
                img = ImageOps.exif_transpose(img)
                reasons.append(f"EXIF rotated (tag={exif_orientation})")

            if file_size > MAX_FILE_SIZE:
                reasons.append(f"too large ({file_size//1024}KB)")
            if img.width > MAX_WIDTH:
                reasons.append(f"too wide ({img.width}px)")

            if not reasons:
                continue

            if fmt == 'PNG':
                img = img.convert('RGBA')
            else:
                img = img.convert('RGB')

            if img.width > MAX_WIDTH:
                ratio = MAX_WIDTH / img.width
                img = img.resize((MAX_WIDTH, int(img.height * ratio)), Image.LANCZOS)

            save_kwargs = {'optimize': True} if fmt == 'PNG' else {'quality': quality}
            img.save(f, fmt, **save_kwargs)
            new_size = f.stat().st_size
            recompressed += 1
            print(f"    {f.name}: {', '.join(reasons)} -> {new_size//1024}KB ({img.width}x{img.height})")
        except Exception as e:
            print(f"    Error checking {f.name}: {e}")

    return recompressed


def recompress_existing():
    """Check all existing ramen and brand images, recompress any that are too large."""
    try:
        from PIL import Image
    except ImportError:
        print("  Pillow not installed, skipping recompression check.")
        return

    # Lowercase all brand logo filenames for consistent matching
    if BRAND_DIR.exists():
        for f in BRAND_DIR.iterdir():
            lower_name = f.name.lower()
            if f.name != lower_name:
                dest = f.with_name(lower_name)
                if not dest.exists():
                    f.rename(dest)
                    print(f"    Renamed {f.name} -> {lower_name}")

    ramen_count = _recompress_dir(IMAGES_DIR, ['*.webp'], 'WEBP', WEBP_QUALITY, Image)
    brand_count = _recompress_dir(BRAND_DIR, ['*.png'], 'PNG', None, Image)

    total = ramen_count + brand_count
    if total:
        print(f"  Recompressed {total} images ({ramen_count} ramen, {brand_count} brand logos)")
    else:
        ramen_files = len(list(IMAGES_DIR.glob('*.webp'))) if IMAGES_DIR.exists() else 0
        brand_files = len(list(BRAND_DIR.glob('*.png')) + list(BRAND_DIR.glob('*.PNG'))) if BRAND_DIR.exists() else 0
        print(f"  All {ramen_files + brand_files} images are properly compressed")


def _fuzzy_rank_ramen(query, ramen_list, limit=50):
    """Fuzzy-match query against brand, variety, country, and style.

    Scoring strategy (0–100 scale):
      - All query words found as substrings       → 85-100 (bonus for coverage)
      - Some query words found as substrings       → 60-84  (proportional to words matched)
      - rapidfuzz partial/token/WRatio fallback     → raw score
      - difflib fallback if no rapidfuzz            → ratio * 100

    Exact review # returns that row only.
    """
    q = (query or "").strip()
    if not q:
        return []
    if q.isdigit():
        rid = int(q)
        for r in ramen_list:
            if r.get("id") == rid:
                return [r]
        return []

    try:
        from rapidfuzz import fuzz
    except ImportError:
        fuzz = None

    q_fold = q.casefold()
    q_words = q_fold.split()

    scored = []
    for r in ramen_list:
        brand = str(r.get("brand") or "")
        variety = str(r.get("variety") or "")
        country = str(r.get("country") or "")
        style = str(r.get("style") or "")
        hay = f"{brand} {variety} {country} {style}"
        hay_fold = hay.casefold()

        # Word-level substring matching
        matched_words = sum(1 for w in q_words if w in hay_fold)
        word_ratio = matched_words / len(q_words) if q_words else 0

        if word_ratio == 1.0:
            coverage = len(q_fold) / len(hay_fold) if hay_fold else 0
            score = 85.0 + 15.0 * coverage
        elif word_ratio > 0:
            score = 60.0 + 24.0 * word_ratio
        elif fuzz:
            score = max(
                fuzz.partial_ratio(q, hay),
                fuzz.token_set_ratio(q, hay),
                fuzz.WRatio(q, hay),
            )
        else:
            from difflib import SequenceMatcher
            score = SequenceMatcher(None, q_fold, hay_fold).ratio() * 100

        scored.append((score, r))

    scored.sort(key=lambda x: (-x[0], x[1].get("id", 0)))
    min_score = 72 if len(q_fold) <= 2 else (55 if fuzz else 40)
    return [r for s, r in scored if s >= min_score][:limit]


def fetch_images_and_popularity(ramen_list, limit=None, panel=None):
    """Fetch product images and popularity scores. Popularity stored in data/popularity.json."""
    IMAGES_DIR.mkdir(exist_ok=True)

    try:
        from PIL import Image
        has_pillow = True
    except ImportError:
        print("  WARNING: Pillow not installed. Saving raw images.")
        has_pillow = False
        Image = None

    pop_map = load_popularity()

    existing = set()
    for f in IMAGES_DIR.glob("*.webp"):
        if f.stem.isdigit():
            existing.add(int(f.stem))

    needed = [r for r in ramen_list
              if r['id'] not in existing
              or r['id'] not in pop_map]

    total_ramen = len(ramen_list)
    missing_pop = sum(1 for r in ramen_list if r['id'] not in pop_map)
    unrefreshed = count_unrefreshed(pop_map)

    if not needed and unrefreshed == 0:
        print(f"Images & popularity: everything up to date ({len(existing)} images).")
        return

    def _priority(r):
        missing_image = r['id'] not in existing
        missing_pop = r['id'] not in pop_map
        tier = 0 if missing_image and missing_pop else 1
        pop_score = -(pop_map.get(r['id'], 0))
        return (tier, pop_score, r['id'])

    needed.sort(key=_priority)

    any_need_popularity = missing_pop > 0 or unrefreshed > 0

    if limit is not None:
        batch = needed[:limit]
        print(f"Processing: {len(needed)} need work -- doing {len(batch)} this run")
    else:
        batch = needed
        print(f"Processing: {len(batch)} ramen need images and/or popularity")
    print(f"  Images:     {len(existing)}/{total_ramen} ({100 * len(existing) / total_ramen:.1f}%)")
    print(f"  Popularity: {total_ramen - missing_pop}/{total_ramen} ({missing_pop} missing, {unrefreshed} to refresh)")

    pw, context, page = None, None, None
    if any_need_popularity:
        try:
            pw, context, page = _create_popularity_browser()
            print("  Launched browser for popularity lookups")
        except Exception as e:
            print(f"  WARNING: Could not launch browser ({e}). Skipping popularity.")
            print("  Run: bash tools/setup.sh")

    downloaded = 0
    scored = 0
    errors = 0
    no_results = 0

    def _get_engine():
        return panel.engine if panel else "google"

    def _do_one(r, progress_prefix):
        nonlocal downloaded, scored, errors, no_results
        out_path = IMAGES_DIR / f"{r['id']}.webp"
        needs_image = not out_path.exists()
        has_image = out_path.exists()
        existing_pop = pop_map.get(r['id'])
        needs_popularity = not existing_pop

        _debug(f"--- #{r['id']} {r['brand']} - {r['variety']}")
        _debug(f"  existing popularity value: {repr(existing_pop)} -> needs_popularity={needs_popularity}")

        todo = []
        if needs_image:
            todo.append("image")
        if needs_popularity:
            todo.append("popularity")

        reason = ", ".join(todo) if todo else "up to date"

        if panel:
            panel.set_progress(f"{progress_prefix} {r['variety'][:30]}")
            panel.set_status(f"Engine: {_get_engine().title()}")

        print(f"    {progress_prefix} #{r['id']} {r['brand']} - {r['variety']}")
        print(f"      Needs: {reason}")

        # --- Image ---
        if needs_image:
            print(f"      Searching Bing for image...")
            query = f'{_quote_brand(r["brand"])} {r["variety"]} the ramen rater'
            candidates = _search_bing(query)

            if not candidates:
                no_results += 1
                print(f"      Image: no candidates found")
            else:
                saved = False
                for j, img_url in enumerate(candidates):
                    print(f"      Image: trying {j+1}/{len(candidates)}: {img_url[:80]}")
                    if _download_image(img_url, out_path, has_pillow, Image):
                        saved = True
                        downloaded += 1
                        has_image = True
                        existing.add(r['id'])
                        _recompress_single(out_path, Image)
                        print(f"      Image: saved")
                        break
                if not saved:
                    errors += 1
                    print(f"      Image: FAILED (all {len(candidates)} candidates failed)")

            time.sleep(0.3)

        # --- Popularity ---
        if needs_popularity and page:
            engine = _get_engine()
            wait = random.uniform(3, 12)
            print(f"      Popularity: searching {engine} (waiting {wait:.0f}s to avoid rate limit)...")
            variety_clean = re.sub(r'\bramen\b', '', r["variety"], flags=re.IGNORECASE).strip()
            variety_clean = re.sub(r'\s+', ' ', variety_clean)
            pop_query = f'{_quote_brand(r["brand"])} {variety_clean} {r.get("country", "")}'.strip()
            time.sleep(wait)
            if engine == "google":
                count = _google_web_result_count(page, pop_query, r["brand"])
            else:
                count = _bing_web_result_count(page, pop_query, r["brand"])
            if count > 0:
                pop_map[r['id']] = count
                scored += 1
                save_popularity(pop_map)
                print(f"      Popularity: {count:,}")
                if panel:
                    panel.record_success(r['id'], r['brand'], r['variety'], count)
                    cur_missing = sum(1 for x in ramen_list if x['id'] not in pop_map)
                    cur_unrefreshed = count_unrefreshed(pop_map)
                    panel.update_scored_total(scored, cur_missing + cur_unrefreshed)
            else:
                no_results += 1
                print(f"      *** POPULARITY ERROR: NO RESULTS for #{r['id']} {r['brand']} — {r['variety']} ***")
                if panel:
                    panel.record_error(r['id'], r['brand'], r['variety'])
        elif existing_pop:
            print(f"      Popularity: {existing_pop:,} (cached)")

        if has_pillow:
            _recompress_scan(Image)

    _ramen_by_id = {r['id']: r for r in ramen_list}

    if panel:
        panel.update_scored_total(0, missing_pop + unrefreshed)

    try:
        idx = 0
        n_batch = len(batch)

        while True:
            if panel:
                if panel.shutting_down:
                    print("\n  Aborted by user.")
                    break
                with panel._lock:
                    req = panel._single_queue.pop(0) if panel._single_queue else None
                    paused = panel._bulk_paused
                if req is not None:
                    found = next((x for x in ramen_list if x.get("id") == req), None)
                    if found:
                        print(f"    [single] #{found['id']} {found['brand']} - {found['variety']}")
                        if found['id'] not in existing and (IMAGES_DIR / f"{found['id']}.webp").exists():
                            existing.add(found['id'])
                        cur_missing = sum(1 for x in ramen_list if x['id'] not in pop_map)
                        cur_unrefreshed = count_unrefreshed(pop_map)
                        _do_one(found, f"[single] pop: {cur_missing} missing, {cur_unrefreshed} to refresh")
                        if panel:
                            panel.set_progress(f"[single] #{found['id']} done")
                        with panel._lock:
                            panel._bulk_paused = False
                        panel._wake.set()
                    else:
                        if panel:
                            panel.set_status(f"No ramen with id {req}")
                        print(f"    [single] no ramen with id {req}")
                    panel._wake.clear()
                    continue

                if paused:
                    if panel._wake.wait(timeout=0.25):
                        panel._wake.clear()
                    continue

            if idx < n_batch:
                r = batch[idx]
                if r['id'] not in existing and (IMAGES_DIR / f"{r['id']}.webp").exists():
                    existing.add(r['id'])
                cur_missing = sum(1 for x in ramen_list if x['id'] not in pop_map)
                cur_unrefreshed = count_unrefreshed(pop_map)
                _do_one(r, f"[{idx + 1}/{n_batch}] pop: {cur_missing} missing, {cur_unrefreshed} to refresh")
                idx += 1
            elif page:
                next_rid = next_refresh_candidate(pop_map)
                if next_rid is None:
                    break
                r = _ramen_by_id.get(next_rid)
                if not r:
                    mark_popularity_refreshed(next_rid)
                    continue
                old_score = pop_map.get(r['id'], 0)
                cur_unrefreshed = count_unrefreshed(pop_map)
                prefix = f"[refresh {cur_unrefreshed} left]"
                print(f"    {prefix} #{r['id']} {r['brand']} — {r['variety']} (current: {old_score:,})")
                if panel:
                    panel.set_progress(f"{prefix} {r['variety'][:30]}")
                    panel.set_status(f"Engine: {_get_engine().title()}")

                if has_pillow:
                    _recompress_scan(Image)

                engine = _get_engine()
                wait = random.uniform(3, 12)
                print(f"      Popularity: searching {engine} (waiting {wait:.0f}s)...")
                variety_clean = re.sub(r'\bramen\b', '', r["variety"], flags=re.IGNORECASE).strip()
                variety_clean = re.sub(r'\s+', ' ', variety_clean)
                pop_query = f'{_quote_brand(r["brand"])} {variety_clean} "ramen" {r.get("country", "")}'.strip()
                time.sleep(wait)
                if engine == "google":
                    count = _google_web_result_count(page, pop_query, r["brand"])
                else:
                    count = _bing_web_result_count(page, pop_query, r["brand"])
                if count > 0:
                    pop_map[r['id']] = count
                    scored += 1
                    save_popularity(pop_map)
                    print(f"      Popularity: {count:,} (was {old_score:,})")
                    if panel:
                        panel.record_success(r['id'], r['brand'], r['variety'], count)
                        cur_missing = sum(1 for x in ramen_list if x['id'] not in pop_map)
                        panel.update_scored_total(scored, cur_missing + count_unrefreshed(pop_map))
                else:
                    no_results += 1
                    print(f"      *** POPULARITY ERROR: NO RESULTS for #{r['id']} {r['brand']} — {r['variety']} (was {old_score:,}) ***")
                    if panel:
                        panel.record_error(r['id'], r['brand'], r['variety'])
                mark_popularity_refreshed(next_rid)
                if panel:
                    cur_missing = sum(1 for x in ramen_list if x['id'] not in pop_map)
                    panel.update_scored_total(scored, cur_missing + count_unrefreshed(pop_map))
            else:
                break
    finally:
        if context:
            context.close()
        if pw:
            pw.stop()

    parts = []
    if downloaded: parts.append(f"{downloaded} images downloaded")
    if scored: parts.append(f"{scored} popularity scores")
    if no_results: parts.append(f"{no_results} no image results")
    if errors: parts.append(f"{errors} image errors")
    print(f"  Done: {', '.join(parts) if parts else 'nothing to do'}")

    recompress_existing()


def main():
    parser = argparse.ArgumentParser(description="Ramen Rater Data Pipeline")
    parser.add_argument('--limit', type=int, default=None, metavar='N',
                        help='Process up to N ramen that still need images/popularity (omit to do all)')
    args = parser.parse_args()

    download_xlsx()

    ramen_list = parse_xlsx()
    if not ramen_list:
        print("No data parsed. Exiting.")
        sys.exit(1)
    save_json(ramen_list)
    cleanup_excluded()

    existing_images = sum(1 for f in IMAGES_DIR.glob("*.webp") if f.stem.isdigit()) if IMAGES_DIR.exists() else 0
    print(f"  {existing_images} ramen already have images")

    print("Checking existing images...")
    recompress_existing()

    # Run scraping with the floating control panel (tkinter needs main thread)
    panel = None
    try:
        panel = ScraperControlPanel(ramen_list)
        print("  Opened scraper control panel (switch engines live)")
        panel.run_with_panel(
            _main_scrape_and_finish, ramen_list, args.limit, panel
        )
    except Exception as e:
        print(f"  Control panel unavailable ({e}), running without GUI")
        _main_scrape_and_finish(ramen_list, args.limit, None)


def _main_scrape_and_finish(ramen_list, limit, panel):
    global _on_captcha
    if panel:
        _on_captcha = panel.record_captcha
    fetch_images_and_popularity(ramen_list, limit=limit, panel=panel)
    _on_captcha = None
    print(f"\nDone! {len(ramen_list)} ramen in database.")


if __name__ == '__main__':
    try:
        main()
    except KeyboardInterrupt:
        print("\nAborted.")
        sys.exit(1)
